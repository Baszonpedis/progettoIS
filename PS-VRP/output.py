import openpyxl as pyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

campi_risultati_euristico=['commessa','macchina', 'data fine stampa', 'minuti setup','minuti processamento','inizio setup','fine setup','inizio lavorazione','fine lavorazione','mt da tagliare','taglio','macchine compatibili','numero coltelli','diametro tubo','veicolo', 'tassativita', 'veicolo tassativo', 'due date (non indicativa)', 'ritardo', 'priorita']

def write_output_soluzione_euristica(schedulazione,nome_file):
    """
    :param schedulazione: lista di dizionari che contengono le informazioni sulle schedule
    :param nome_file: percorso che indica dove salvare il file e con che nome
    :return: file excel in cui vado a fare la print delle informazioni relative alla schedulazione
    """
    wb=pyxl.Workbook() #inizializzo il file
    ws1=wb.active #si prende il foglio attivo
    ws1.title='Schedulazione' #si rinomina il titolo del foglio
    ws1.append(campi_risultati_euristico) #vado ad inserire il nome delle colonne
    ws1.column_dimensions['A'].width=10 #si settano le dimensioni delle colonne
    ws1.column_dimensions['B'].width=10
    ws1.column_dimensions['C'].width=15
    ws1.column_dimensions['D'].width=15
    ws1.column_dimensions['E'].width=15
    ws1.column_dimensions['F'].width=25
    ws1.column_dimensions['G'].width=25
    ws1.column_dimensions['H'].width=25
    ws1.column_dimensions['I'].width=15
    ws1.column_dimensions['J'].width=15
    ws1.column_dimensions['K'].width=15
    ws1.column_dimensions['L'].width=50
    ws1.column_dimensions['M'].width=15
    ws1.column_dimensions['N'].width=15
    ws1.column_dimensions['O'].width=10
    ws1.column_dimensions['P'].width=10
    ws1.column_dimensions['Q'].width=15
    ws1.column_dimensions['R'].width=15

    start_row=2 #inizializzo la riga in cui andrò a printare. si parte dalla seconda in quanto la prima è occupata dai titoli
    start_column=1 #inizializzo le colonne in cui andrò a printare si parte dalla prima e si andrà avanti fino all'ultimo campo
    for schedula in schedulazione: #per ogni dizionario nella lista
        for chiave,valore in schedula.items(): #vado a prendere tutti i valori all'interno del dizionario
            if chiave=='macchine compatibili':
                valore.sort()
                valore=" ;".join(valore)
            if type(valore)==pd.Timestamp: #se sto considerando un campo contenente una data pandas devo convertirla
                valore=valore.strftime("%d-%m-%Y %H:%M:%S") #converto in data (giorno-mese-anno)
            if chiave=='veicolo' and valore != None and not isinstance(valore, str):
                valore = valore.nome
            ws1.cell(row=start_row,column=start_column,value=valore) #assegno il valore in questione alla cella
            start_column+=1 #se non sono finiti i campi avanzo di una colonna
        start_row+=1 #quando sono finiti i campi passo alla riga successiva
        start_column=1 #quando sono finiti i campi riparto dalla prima colonna
    wb.save(nome_file) #salvo il file excel con il nome che passo come parametro

def write_output_ridotto(schedulazione,nome_file):
    """
    :param schedulazione: lista di dizionari con info sulla schedulazione
    :param nome_file: percorso file excel
    :return: file excel abbellito
    """
    campi_risultati_ridotti = ['commessa', 'macchina', 'inizio_setup', 'inizio_lavorazione', 'tassativita', 'priorita']

    # Colori pastello tenui (HEX)
    colori_pastello = [
        'CCE5FF',  # azzurro chiaro
        'D5E8D4',  # verde pallido
        'FCE5CD',  # arancio chiarissimo
        'EAD1DC',  # rosa tenue
        'FFF2CC',  # giallo chiaro
        'D9D2E9',  # lilla chiaro
        'E2EFDA',  # verde menta
        'F4CCCC',  # rosato
    ]

    # Costruisci mappa macchina → colore
    macchine = list({schedula['macchina'] for schedula in schedulazione})
    macchina_colori = {
        macchina: colori_pastello[i % len(colori_pastello)] for i, macchina in enumerate(macchine)
    }

    # Workbook e foglio
    wb = pyxl.Workbook()
    ws = wb.active
    ws.title = 'Schedulazione'

    # Stili base
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Scrivi intestazioni
    for idx, campo in enumerate(campi_risultati_ridotti, start=1):
        cell = ws.cell(row=1, column=idx, value=campo)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Scrivi dati con colore
    for row_idx, schedula in enumerate(schedulazione, start=2):
        macchina = schedula.get('macchina')
        colore = macchina_colori.get(macchina, 'FFFFFF')
        fill = PatternFill(start_color=colore, end_color=colore, fill_type='solid')

        for col_idx, chiave in enumerate(campi_risultati_ridotti, start=1):
            valore = schedula.get(chiave, '')
            if isinstance(valore, pd.Timestamp):
                valore = valore.strftime("%d-%m-%Y %H:%M:%S")
            cell = ws.cell(row=row_idx, column=col_idx, value=valore)
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = fill

    # Colonne larghezza base
    col_widths = [15, 15, 22, 22]
    for idx, width in zip(range(1, len(campi_risultati_ridotti) + 1), col_widths):
        ws.column_dimensions[pyxl.utils.get_column_letter(idx)].width = width

    # Congela intestazione
    ws.freeze_panes = 'A2'

    # Salva file
    wb.save(nome_file)

def write_output_ridotto_txt(schedulazione, nome_file):
    """
    Salva la schedulazione in un .txt formattato a tabella leggibile
    """
    campi_risultati_ridotti = ['commessa', 'macchina', 'inizio_setup', 'inizio_lavorazione', 'tassativita', 'priorita']

    # Larghezza colonne di base
    larghezze = [15, 15, 22, 22]  # Ho allargato un po' anche le date

    with open(nome_file, 'w', encoding='utf-8') as f:
        # Linea intestazione
        intestazione = ''
        for i, (campo, larghezza) in enumerate(zip(campi_risultati_ridotti, larghezze)):
            intestazione += f"{campo:<{larghezza}}"
            # Aggiungi separazione extra solo tra terza e quarta colonna
            if i == 2:
                intestazione += '   '  # 3 spazi extra
        f.write(intestazione + '\n')

        # Riga di separazione
        separatore = ''
        for i, larghezza in enumerate(larghezze):
            separatore += '-' * larghezza
            if i == 2:
                separatore += '   '
        f.write(separatore + '\n')

        # Dati
        for schedula in schedulazione:
            riga = ''
            for i, (chiave, larghezza) in enumerate(zip(campi_risultati_ridotti, larghezze)):
                valore = schedula.get(chiave, '')
                if isinstance(valore, pd.Timestamp):
                    valore = valore.strftime("%d-%m-%Y %H:%M:%S")
                riga += f"{str(valore):<{larghezza}}"
                if i == 2:
                    riga += '   '
            f.write(riga + '\n')

def write_error_output(df, nome_file):
    """
    :param df: dataframe contenente tutte le commesse e tutti i campi
    :param nome_file: percorso che indica dove salvare il file e con che nome
    :return: file excel in cui vado a fare la print delle commesse e i relativi campi vuoti
    """
    # Nomi dei campi "speciali"
    FLAG_TASSATIVO_FIELD = "flag tassativo taglio per schedulatore"
    ID_SPEDIZIONE_FIELD = "id spedizione"
    CODICE_DI_ZONA_FIELD = "Commesse::CODICE DI ZONA"

    # Assicurati che i campi esistano nel DataFrame
    required_special_fields = [FLAG_TASSATIVO_FIELD, ID_SPEDIZIONE_FIELD, CODICE_DI_ZONA_FIELD]
    for field in required_special_fields:
        if field not in df.columns:
            print(f"Attenzione: Il campo '{field}' non è presente nel DataFrame. Verrà ignorato nella logica speciale.")
            # Puoi scegliere di aggiungere il campo con NaN se vuoi che la logica funzioni
            # df[field] = pd.NA


    wb = pyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Error'
    nomi_colonne = list(df.columns)
    
    # Inizializza un DataFrame per le commesse da includere nel file di output
    df_to_output = pd.DataFrame(columns=nomi_colonne)

    # Identifica le commesse che devono essere incluse nel file
    commesse_da_includere_indices = []

    for riga_df_index, row_data in df.iterrows():
        # Crea una copia temporanea della riga per la valutazione dei campi vuoti rilevanti
        row_for_evaluation = row_data.copy()

        # Logica speciale per 'flag tassativo taglio per schedulatore' e 'id spedizione'
        # Se entrambi sono nulli, li consideriamo "non nulli" per la valutazione complessiva della riga
        if (FLAG_TASSATIVO_FIELD in row_for_evaluation and ID_SPEDIZIONE_FIELD in row_for_evaluation and
            pd.isna(row_for_evaluation[FLAG_TASSATIVO_FIELD]) and
            pd.isna(row_for_evaluation[ID_SPEDIZIONE_FIELD])):
            row_for_evaluation[FLAG_TASSATIVO_FIELD] = "BOTH_NULL_OK"
            row_for_evaluation[ID_SPEDIZIONE_FIELD] = "BOTH_NULL_OK"

        # Tratta 'codice di zona' come non essenziale
        if CODICE_DI_ZONA_FIELD in row_for_evaluation and pd.isna(row_for_evaluation[CODICE_DI_ZONA_FIELD]):
            row_for_evaluation[CODICE_DI_ZONA_FIELD] = "NOT_ESSENTIAL_NULL"

        # Controlla se ci sono ancora campi nulli "rilevanti" nella riga
        if row_for_evaluation.isnull().any():
            commesse_da_includere_indices.append(riga_df_index)

    # Filtra il DataFrame originale per includere solo le commesse con campi vuoti rilevanti
    df_filtered_for_output = df.loc[commesse_da_includere_indices].copy()

    # Se non ci sono commesse con campi vuoti rilevanti, salva un file vuoto o con solo intestazioni
    if df_filtered_for_output.empty:
        ws1.append(nomi_colonne)
        wb.save(nome_file)
        print(f"Nessuna commessa con campi vuoti rilevanti trovata. Il file '{nome_file}' è stato creato con sole intestazioni.")
        return

    # Aggiungi le intestazioni al foglio Excel
    ws1.append(nomi_colonne)

    # Impostazione delle dimensioni delle colonne (basato sul df originale per avere tutte le colonne)
    for i, col_name in enumerate(nomi_colonne):
        ws1.column_dimensions[chr(65 + i)].width = 30 # A=65, B=66, etc.
    # Puoi aggiungere eccezioni specifiche per colonne come prima
    # ws1.column_dimensions['M'].width = 35 # Esempio

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Ora iteriamo sul DataFrame filtrato per scrivere e evidenziare
    # Manteniamo un contatore di riga per Excel
    excel_row_counter = 2 # Inizia dalla riga 2 dopo le intestazioni

    for riga_df_index, row_data in df_filtered_for_output.iterrows():
        # Creiamo una versione della riga con 'CAMPO VUOTO' per la stampa
        row_for_printing = row_data.fillna('CAMPO VUOTO')

        for col_index, col_name in enumerate(nomi_colonne):
            value_to_print = row_for_printing[col_name]
            ws1.cell(row=excel_row_counter, column=col_index + 1, value=value_to_print)

            original_value = df.at[riga_df_index, col_name]

            # Logica di evidenziazione
            if pd.isna(original_value): # Se il campo originale era nullo
                # Non evidenziare 'codice di zona' se è nullo
                if col_name == CODICE_DI_ZONA_FIELD:
                    continue # Passa al prossimo campo senza evidenziare

                # Non evidenziare 'flag tassativo taglio per schedulatore' o 'id spedizione'
                # se entrambi erano nulli nel DF originale per questa commessa
                if (col_name == FLAG_TASSATIVO_FIELD or col_name == ID_SPEDIZIONE_FIELD):
                    if (pd.isna(df.at[riga_df_index, FLAG_TASSATIVO_FIELD]) and
                        pd.isna(df.at[riga_df_index, ID_SPEDIZIONE_FIELD])):
                        continue # Entrambi nulli, non evidenziare nessuno dei due
                    else:
                        # Uno solo è nullo, l'altro no (o non esiste), quindi evidenzia
                        ws1.cell(row=excel_row_counter, column=col_index + 1).fill = fill
                else:
                    # Per tutti gli altri campi nulli, evidenzia
                    ws1.cell(row=excel_row_counter, column=col_index + 1).fill = fill
        
        excel_row_counter += 1

    wb.save(nome_file)

def write_veicoli_error_output(df, nome_file):
    wb = pyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Errori Veicoli'
    
    # Titoli colonne
    nomi_colonne = list(df.columns)
    ws1.append(nomi_colonne)
    
    # Imposta larghezza colonne dinamicamente
    for i, col in enumerate(nomi_colonne, start=1):
        ws1.column_dimensions[chr(64 + i)].width = 30
    
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    start_row = 2
    df = df.fillna("CAMPO VUOTO")
    
    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            cell = ws1.cell(row=start_row, column=col_index + 1, value=value)
            if value == "CAMPO VUOTO":
                cell.fill = fill
        start_row += 1
    
    wb.save(nome_file)


def write_tassative_error_output(df, nome_file, nome_foglio="Errori"):
    """
    Scrive i dati del DataFrame in un foglio Excel con evidenziazione dei campi vuoti.
    Se il file esiste già, aggiunge un foglio o sovrascrive quello con lo stesso nome.
    """

    # Se il file esiste già → aprilo, altrimenti creane uno nuovo
    try:
        wb = pyxl.load_workbook(nome_file)
    except FileNotFoundError:
        wb = pyxl.Workbook()
        # rimuovi foglio vuoto di default
        default_sheet = wb.active
        wb.remove(default_sheet)

    # Se il foglio esiste già → rimuovilo (sovrascrittura)
    if nome_foglio in wb.sheetnames:
        std = wb[nome_foglio]
        wb.remove(std)

    # Crea nuovo foglio
    ws1 = wb.create_sheet(title=nome_foglio)

    # Titoli colonne
    nomi_colonne = list(df.columns)
    ws1.append(nomi_colonne)
    
    # Imposta larghezza colonne dinamicamente
    for i, _ in enumerate(nomi_colonne, start=1):
        col_letter = get_column_letter(i)
        ws1.column_dimensions[col_letter].width = 30
    
    # Evidenziazione gialla
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    start_row = 2
    df = df.fillna("CAMPO VUOTO")
    
    for _, row in df.iterrows():
        for col_index, value in enumerate(row, start=1):
            cell = ws1.cell(row=start_row, column=col_index, value=value)
            if value == "CAMPO VUOTO":
                cell.fill = fill
        start_row += 1
    
    # Salva file
    wb.save(nome_file)
    print(f"Foglio '{nome_foglio}' scritto in '{nome_file}' (sovrascrittura se già esisteva).")
